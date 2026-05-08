from __future__ import annotations

import re
import zipfile
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree


class DocumentConversionError(RuntimeError):
    pass


def convert_source_file_to_markdown(source_path: Path, processor_type: str) -> str:
    normalized_processor = str(processor_type or "").strip().lower()
    if normalized_processor == "docx_markdown":
        return convert_docx_to_markdown(source_path)
    if normalized_processor == "excel_markdown":
        return convert_excel_to_markdown(source_path)
    raise DocumentConversionError(f"Unsupported document processor: {processor_type}")


def convert_docx_to_markdown(source_path: Path) -> str:
    try:
        markdown = _convert_docx_with_python_docx(source_path)
    except Exception:
        markdown = _convert_docx_with_ooxml(source_path)
    markdown = _clean_markdown(markdown)
    if not markdown:
        raise DocumentConversionError("Word 文件未提取到可入库文本")
    return markdown


def convert_excel_to_markdown(source_path: Path) -> str:
    try:
        markdown = _convert_excel_with_openpyxl(source_path)
    except Exception:
        markdown = _convert_excel_with_ooxml(source_path)
    markdown = _clean_markdown(markdown)
    if not markdown:
        raise DocumentConversionError("Excel 文件未提取到可入库文本")
    return markdown


def _convert_docx_with_python_docx(source_path: Path) -> str:
    from docx import Document
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    document = Document(str(source_path))
    parts: list[str] = []
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            markdown = _docx_paragraph_to_markdown(Paragraph(child, document))
            if markdown:
                parts.append(markdown)
        elif isinstance(child, CT_Tbl):
            markdown = _rows_to_markdown_table(_docx_table_rows(Table(child, document)))
            if markdown:
                parts.append(markdown)
    return "\n\n".join(parts)


def _docx_paragraph_to_markdown(paragraph: Any) -> str:
    text = _normalize_inline_text(paragraph.text)
    if not text:
        return ""
    style_name = ""
    try:
        style_name = str(paragraph.style.name or "").strip().lower()
    except Exception:
        style_name = ""

    heading_match = re.search(r"heading\s+([1-6])", style_name)
    if heading_match:
        level = int(heading_match.group(1))
        return f"{'#' * level} {text}"
    if style_name == "title":
        return f"# {text}"
    if "list bullet" in style_name or style_name.startswith("bullet"):
        return f"- {text}"
    if "list number" in style_name or style_name.startswith("number"):
        return f"1. {text}"
    return text


def _docx_table_rows(table: Any) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in table.rows:
        values: list[str] = []
        for cell in row.cells:
            values.append(_normalize_multiline_text(cell.text))
        rows.append(values)
    return rows


def _convert_docx_with_ooxml(source_path: Path) -> str:
    with zipfile.ZipFile(source_path) as archive:
        document_xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(document_xml)
    body = next((item for item in root.iter() if _local_name(item.tag) == "body"), None)
    if body is None:
        return ""

    parts: list[str] = []
    for child in list(body):
        local = _local_name(child.tag)
        if local == "p":
            text = _normalize_inline_text(_collect_xml_text(child))
            if text:
                parts.append(text)
        elif local == "tbl":
            rows: list[list[str]] = []
            for row in child.iter():
                if _local_name(row.tag) != "tr":
                    continue
                cells = [
                    _normalize_multiline_text(_collect_xml_text(cell))
                    for cell in list(row)
                    if _local_name(cell.tag) == "tc"
                ]
                if cells:
                    rows.append(cells)
            table_markdown = _rows_to_markdown_table(rows)
            if table_markdown:
                parts.append(table_markdown)
    return "\n\n".join(parts)


def _convert_excel_with_openpyxl(source_path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(str(source_path), read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for worksheet in workbook.worksheets:
            rows = [
                [_format_cell_value(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            sheet_markdown = _sheet_rows_to_markdown(worksheet.title, rows)
            if sheet_markdown:
                parts.append(sheet_markdown)
        return "\n\n".join(parts)
    finally:
        workbook.close()


def _convert_excel_with_ooxml(source_path: Path) -> str:
    with zipfile.ZipFile(source_path) as archive:
        shared_strings = _read_xlsx_shared_strings(archive)
        sheet_targets = _read_xlsx_sheet_targets(archive)
        parts: list[str] = []
        for sheet_name, sheet_path in sheet_targets:
            try:
                sheet_xml = archive.read(sheet_path)
            except KeyError:
                continue
            rows = _read_xlsx_sheet_rows(sheet_xml, shared_strings)
            sheet_markdown = _sheet_rows_to_markdown(sheet_name, rows)
            if sheet_markdown:
                parts.append(sheet_markdown)
        return "\n\n".join(parts)


def _read_xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        raw_xml = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ElementTree.fromstring(raw_xml)
    strings: list[str] = []
    for item in root:
        if _local_name(item.tag) != "si":
            continue
        strings.append("".join(node.text or "" for node in item.iter() if _local_name(node.tag) == "t"))
    return strings


def _read_xlsx_sheet_targets(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    relationships = _read_xlsx_workbook_relationships(archive)
    sheets: list[tuple[str, str]] = []
    for sheet in workbook.iter():
        if _local_name(sheet.tag) != "sheet":
            continue
        name = str(sheet.attrib.get("name") or "").strip() or "Sheet"
        relationship_id = (
            sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            or sheet.attrib.get("id")
            or ""
        )
        target = relationships.get(relationship_id)
        if not target:
            sheet_id = str(sheet.attrib.get("sheetId") or "").strip()
            target = f"xl/worksheets/sheet{sheet_id}.xml" if sheet_id else ""
        if target:
            sheets.append((name, target))
    return sheets


def _read_xlsx_workbook_relationships(archive: zipfile.ZipFile) -> dict[str, str]:
    try:
        raw_xml = archive.read("xl/_rels/workbook.xml.rels")
    except KeyError:
        return {}
    root = ElementTree.fromstring(raw_xml)
    relationships: dict[str, str] = {}
    for relationship in root:
        relationship_id = str(relationship.attrib.get("Id") or "").strip()
        target = str(relationship.attrib.get("Target") or "").strip()
        if not relationship_id or not target:
            continue
        if target.startswith("/"):
            normalized = target.lstrip("/")
        else:
            normalized = str(PurePosixPath("xl") / target)
        relationships[relationship_id] = normalized
    return relationships


def _read_xlsx_sheet_rows(sheet_xml: bytes, shared_strings: list[str]) -> list[list[str]]:
    root = ElementTree.fromstring(sheet_xml)
    rows: list[list[str]] = []
    for row in root.iter():
        if _local_name(row.tag) != "row":
            continue
        values_by_column: dict[int, str] = {}
        next_column = 0
        for cell in list(row):
            if _local_name(cell.tag) != "c":
                continue
            column_index = _column_index_from_reference(str(cell.attrib.get("r") or ""))
            if column_index is None:
                column_index = next_column
            values_by_column[column_index] = _read_xlsx_cell_value(cell, shared_strings)
            next_column = max(next_column, column_index + 1)
        if values_by_column:
            max_column = max(values_by_column)
            rows.append([values_by_column.get(index, "") for index in range(max_column + 1)])
    return rows


def _read_xlsx_cell_value(cell: ElementTree.Element, shared_strings: list[str]) -> str:
    cell_type = str(cell.attrib.get("t") or "").strip()
    if cell_type == "inlineStr":
        return _normalize_multiline_text(_collect_xml_text(cell))

    value_node = next((item for item in cell if _local_name(item.tag) == "v"), None)
    raw_value = value_node.text if value_node is not None else ""
    if raw_value is None:
        raw_value = ""
    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)]
        except (IndexError, TypeError, ValueError):
            return raw_value
    if cell_type == "b":
        return "TRUE" if raw_value == "1" else "FALSE"
    return _normalize_multiline_text(raw_value)


def _sheet_rows_to_markdown(sheet_name: str, rows: list[list[str]]) -> str:
    trimmed_rows = _trim_table(rows)
    if not trimmed_rows:
        return ""

    first_row_index = next(
        (index for index, row in enumerate(trimmed_rows) if any(_has_text(value) for value in row)),
        None,
    )
    if first_row_index is None:
        return ""

    header_row = trimmed_rows[first_row_index]
    data_rows = trimmed_rows[first_row_index + 1 :]
    headers = _normalize_headers(header_row)
    table = _rows_to_markdown_table([headers, *data_rows])
    if not table:
        return ""
    return f"## 工作表：{sheet_name}\n\n{table}"


def _rows_to_markdown_table(rows: list[list[str]]) -> str:
    trimmed_rows = _trim_table(rows)
    if not trimmed_rows:
        return ""
    width = max(len(row) for row in trimmed_rows)
    normalized_rows = [row + [""] * (width - len(row)) for row in trimmed_rows]
    headers = _normalize_headers(normalized_rows[0])
    data_rows = normalized_rows[1:] or [[""] * width]

    lines = [
        "| " + " | ".join(_escape_markdown_cell(value) for value in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in data_rows:
        lines.append("| " + " | ".join(_escape_markdown_cell(value) for value in row) + " |")
    return "\n".join(lines)


def _normalize_headers(values: list[str]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for index, value in enumerate(values):
        text = _normalize_inline_text(value)
        if not text or text in seen:
            text = f"Column {_column_name(index)}"
        seen.add(text)
        headers.append(text)
    if not headers:
        return ["Column A"]
    return headers


def _trim_table(rows: list[list[Any]]) -> list[list[str]]:
    normalized_rows = [[_format_cell_value(value) for value in row] for row in rows]
    while normalized_rows and not any(_has_text(value) for value in normalized_rows[-1]):
        normalized_rows.pop()
    if not normalized_rows:
        return []

    max_column = -1
    for row in normalized_rows:
        for index, value in enumerate(row):
            if _has_text(value):
                max_column = max(max_column, index)
    if max_column < 0:
        return []
    return [row[: max_column + 1] + [""] * max(0, max_column + 1 - len(row)) for row in normalized_rows]


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _normalize_multiline_text(str(value))


def _escape_markdown_cell(value: Any) -> str:
    text = _normalize_multiline_text(value)
    return text.replace("\\", "\\\\").replace("|", "\\|")


def _normalize_inline_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _normalize_multiline_text(value: Any) -> str:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    lines = [_normalize_inline_text(line) for line in text.split("\n")]
    return "<br>".join(line for line in lines if line)


def _clean_markdown(markdown: str) -> str:
    lines = [line.rstrip() for line in str(markdown or "").splitlines()]
    text = "\n".join(lines)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _collect_xml_text(element: ElementTree.Element) -> str:
    return "".join(node.text or "" for node in element.iter() if _local_name(node.tag) == "t")


def _column_index_from_reference(reference: str) -> int | None:
    match = re.match(r"([A-Za-z]+)", reference)
    if not match:
        return None
    index = 0
    for char in match.group(1).upper():
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _column_name(index: int) -> str:
    value = index + 1
    name = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        name = chr(ord("A") + remainder) + name
    return name or "A"


def _has_text(value: Any) -> bool:
    return bool(str(value or "").strip())


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag
